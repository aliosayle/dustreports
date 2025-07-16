from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import pyodbc
import numpy as np
from datetime import datetime
import io
import threading
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# Try to import interbase for direct InterBase connection
try:
    import interbase
    INTERBASE_AVAILABLE = True
    print("✅ InterBase Python driver available for direct connection")
except ImportError:
    INTERBASE_AVAILABLE = False
    print("⚠️ InterBase Python driver not available, will use ODBC as fallback")

# Suppress warnings to match notebook behavior
warnings.filterwarnings('ignore')

app = Flask(__name__)
app.config['SECRET_KEY'] = 'iba-dust-reports-2025'

# Database connection parameters
DATABASE_CONFIG = {
    'DATA_SOURCE': "100.200.2.1",
    'DATABASE_PATH': r"D:\dolly2008\fer2015.dol",
    'USERNAME': "ALIOSS",
    'PASSWORD': "Ali@123",
    'CLIENT_LIBRARY': r"C:\Users\User\Downloads\Compressed\ibclient64-14.1_x86-64\ibclient64-14.1.dll"  # InterBase client library location
}

# ODBC connection string (fallback)
connection_string = (
    f"DRIVER=Devart ODBC Driver for InterBase;"
    f"Data Source={DATABASE_CONFIG['DATA_SOURCE']};"
    f"Database={DATABASE_CONFIG['DATABASE_PATH']};"
    f"User ID={DATABASE_CONFIG['USERNAME']};"
    f"Password={DATABASE_CONFIG['PASSWORD']};"
)

# Global cache for dataframes (matching notebook structure)
dataframes = {}
cache_lock = threading.Lock()
cache_loading = False

def connect_and_load_table(table_name):
    """Load a table from the database using direct InterBase connection only"""
    try:
        print(f"🔄 Connecting to database for table {table_name}...")
        
        # Use direct InterBase connection only (no fallback to ODBC)
        if not INTERBASE_AVAILABLE:
            raise Exception("InterBase Python library not available")
        
        print(f"🔗 Using direct InterBase connection for {table_name}...")
        
        # Build direct connection for InterBase
        # Format: host:database_path
        dsn = f"{DATABASE_CONFIG['DATA_SOURCE']}:{DATABASE_CONFIG['DATABASE_PATH']}"
        print(f"📡 DSN: {dsn}")
        print(f"📚 Client Library: {DATABASE_CONFIG.get('CLIENT_LIBRARY', 'system default')}")
        
        # Connect with explicit client library if specified
        if 'CLIENT_LIBRARY' in DATABASE_CONFIG:
            conn = interbase.connect(
                dsn=dsn,
                user=DATABASE_CONFIG['USERNAME'],
                password=DATABASE_CONFIG['PASSWORD'],
                ib_library_name=DATABASE_CONFIG['CLIENT_LIBRARY'],
                charset='NONE'  # Use UTF-8 charset for better character compatibility
            )
        else:
            conn = interbase.connect(
                dsn=dsn,
                user=DATABASE_CONFIG['USERNAME'],
                password=DATABASE_CONFIG['PASSWORD'],
                charset='NONE'  # Use UTF-8 charset for better character compatibility
            )
        
        print(f"✅ Direct InterBase connection successful for {table_name}")
        
        # Execute query and fetch data
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM {table_name}")
        
        # Get column names
        columns = [desc[0] for desc in cursor.description]
        
        # Fetch all rows
        rows = cursor.fetchall()
        
        # Convert to DataFrame
        df = pd.DataFrame(rows, columns=columns)
        
        conn.close()
        print(f"✅ {table_name}: {df.shape[0]:,} rows × {df.shape[1]} columns (direct connection)")
        return df
        
    except Exception as e:
        print(f"❌ {table_name}: Failed to load - {e}")
        print(f"   Error type: {type(e).__name__}")
        print(f"   DSN attempted: {DATABASE_CONFIG['DATA_SOURCE']}:{DATABASE_CONFIG['DATABASE_PATH']}")
        print(f"   Client Library: {DATABASE_CONFIG.get('CLIENT_LIBRARY', 'system default')}")
        return None

def load_dataframes():
    """Load all tables with descriptive names (matching notebook exactly)"""
    global dataframes, cache_loading
    
    try:
        print("Loading database tables...")
        
        # Test connection first
        try:
            if INTERBASE_AVAILABLE:
                try:
                    print("🔗 Testing direct InterBase connection...")
                    dsn = f"{DATABASE_CONFIG['DATA_SOURCE']}:{DATABASE_CONFIG['DATABASE_PATH']}"
                    print(f"📡 Connecting to DSN: {dsn}")
                    print(f"👤 User: {DATABASE_CONFIG['USERNAME']}")
                    print(f"📚 Client Library: {DATABASE_CONFIG.get('CLIENT_LIBRARY', 'system default')}")
                    
                    # Connect with explicit client library if specified
                    if 'CLIENT_LIBRARY' in DATABASE_CONFIG:
                        test_conn = interbase.connect(
                            dsn=dsn,
                            user=DATABASE_CONFIG['USERNAME'],
                            password=DATABASE_CONFIG['PASSWORD'],
                            ib_library_name=DATABASE_CONFIG['CLIENT_LIBRARY'],
                            charset='UTF8'  # Use UTF-8 charset for better character compatibility
                        )
                    else:
                        test_conn = interbase.connect(
                            dsn=dsn,
                            user=DATABASE_CONFIG['USERNAME'],
                            password=DATABASE_CONFIG['PASSWORD'],
                            charset='UTF8'  # Use UTF-8 charset for better character compatibility
                        )
                    test_conn.close()
                    print("✅ Direct InterBase connection test successful")
                    print("🚀 Will use direct InterBase connection for all tables")
                except Exception as interbase_error:
                    print(f"⚠️ Direct InterBase connection test failed: {interbase_error}")
                    print(f"� Error type: {type(interbase_error).__name__}")
                    print("❌ WILL NOT fallback to ODBC - fixing direct connection...")
                    raise Exception(f"Direct InterBase connection failed: {interbase_error}")
            else:
                print("❌ InterBase library not available")
                raise Exception("InterBase library not available")
        except Exception as e:
            print(f"❌ Database connection test failed: {e}")
            cache_loading = False
            raise Exception(f"Database connection failed: {e}")
        
        sites_df = connect_and_load_table('ALLSTOCK')          # Site/Location master data
        categories_df = connect_and_load_table('DETDESCR')     # Category definitions  
        invoice_headers_df = connect_and_load_table('INVOICE') # Invoice headers
        sales_details_df = connect_and_load_table('ITEMS')     # Sales transaction details
        vouchers_df = connect_and_load_table('PAYM')           # Payment vouchers
        accounts_df = connect_and_load_table('SACCOUNT')       # Statement of accounts
        inventory_items_df = connect_and_load_table('STOCK')   # Items/Products master
        inventory_transactions_df = connect_and_load_table('ALLITEM') # All inventory transactions
        
        # Create dataframes dictionary with descriptive names (matching notebook exactly)
        temp_dataframes = {
            'sites': sites_df,
            'categories': categories_df, 
            'invoice_headers': invoice_headers_df,
            'sales_details': sales_details_df,
            'vouchers': vouchers_df,
            'accounts': accounts_df,
            'inventory_items': inventory_items_df,
            'inventory_transactions': inventory_transactions_df
        }
        
        # Remove None values and show summary (matching notebook exactly)
        dataframes = {k: v for k, v in temp_dataframes.items() if v is not None}
        print(f"\n✅ Successfully loaded {len(dataframes)} tables:")
        for name, df in dataframes.items():
            print(f"  {name}: {df.shape[0]:,} rows × {df.shape[1]} columns")
        
        if len(dataframes) == 0:
            raise Exception("No tables were loaded successfully")
            
        cache_loading = False
        return dataframes
        
    except Exception as e:
        print(f"❌ Error in load_dataframes: {e}")
        cache_loading = False
        raise e

def calculate_stock_and_sales(item_code=None, site_code=None, from_date=None, to_date=None, as_of_date=None, site_codes=None, category_id=None):
    """
    Calculate current stock and sales (matching notebook exactly)
    """
    if not dataframes:
        return None
        
    # Start with inventory transactions for stock calculation
    df_stock = dataframes['inventory_transactions'].copy()
    
    # Filter by date if specified (show stock as of a specific date)
    if as_of_date:
        as_of_dt = pd.to_datetime(as_of_date)
        if 'FDATE' in df_stock.columns:
            df_stock['FDATE'] = pd.to_datetime(df_stock['FDATE'], errors='coerce')
            df_stock = df_stock[df_stock['FDATE'] <= as_of_dt]
            print(f"📅 Filtering stock transactions up to {as_of_date} - {len(df_stock)} transactions found")
        else:
            print("⚠️ FDATE column not found in inventory transactions - cannot filter by date")
    
    # Filter by item if specified
    if item_code:
        df_stock = df_stock[df_stock['ITEM'] == item_code]
        if df_stock.empty:
            print(f"❌ No stock transactions found for item: {item_code}")
            return None
    
    # Filter by site if specified (single site - legacy support)
    if site_code:
        df_stock = df_stock[df_stock['SITE'] == site_code]
        if df_stock.empty:
            print(f"❌ No stock transactions found for site: {site_code}")
            return None
    
    # Filter by multiple sites if specified
    if site_codes and isinstance(site_codes, list) and len(site_codes) > 0:
        df_stock = df_stock[df_stock['SITE'].isin(site_codes)]
        if df_stock.empty:
            print(f"❌ No stock transactions found for sites: {site_codes}")
            return None
        print(f"📍 Filtering for {len(site_codes)} sites: {site_codes}")
    
    # Filter by category if specified
    if category_id:
        # First, get items in the specified category
        if 'inventory_items' in dataframes and dataframes['inventory_items'] is not None:
            items_in_category = dataframes['inventory_items'][
                dataframes['inventory_items']['CATEGORY'].astype(str) == str(category_id)
            ]['ITEM'].unique()
            
            if len(items_in_category) == 0:
                print(f"❌ No items found in category: {category_id}")
                return None
            
            # Filter stock transactions to only include items in the category
            df_stock = df_stock[df_stock['ITEM'].isin(items_in_category)]
            if df_stock.empty:
                print(f"❌ No stock transactions found for items in category: {category_id}")
                return None
            print(f"📋 Filtering for category {category_id}: {len(items_in_category)} items")
        else:
            print("⚠️ Inventory items data not available for category filtering")
            return None
    
    # Fill NaN values with 0 for calculations
    df_stock['DEBITQTY'] = df_stock['DEBITQTY'].fillna(0)
    df_stock['CREDITQTY'] = df_stock['CREDITQTY'].fillna(0)
    
    # Calculate stock by grouping
    if item_code and site_code:
        # Single item, single site
        result_df = pd.DataFrame({
            'SITE': [site_code],
            'ITEM': [item_code],
            'TOTAL_IN': [df_stock['DEBITQTY'].sum()],
            'TOTAL_OUT': [df_stock['CREDITQTY'].sum()],
            'CURRENT_STOCK': [df_stock['DEBITQTY'].sum() - df_stock['CREDITQTY'].sum()],
            'STOCK_TRANSACTIONS': [len(df_stock)],
            'SELECTED_SITES': ['']
        })
    elif item_code:
        # Single item, all sites (or selected sites)
        if site_codes and isinstance(site_codes, list) and len(site_codes) > 0:
            # Multiple sites selected - aggregate across all selected sites for this item
            # Get site names for display
            site_names_display = "Multiple Sites"
            if 'sites' in dataframes and dataframes['sites'] is not None:
                site_name_map = dict(zip(dataframes['sites']['ID'], dataframes['sites']['SITE']))
                selected_site_names = [site_name_map.get(site_id, site_id) for site_id in site_codes]
                site_names_display = f"{', '.join(selected_site_names[:3])}" + (f" (+{len(selected_site_names)-3} more)" if len(selected_site_names) > 3 else "")
            
            result_df = pd.DataFrame({
                'SITE': [f"Multiple Sites ({len(site_codes)})"],
                'ITEM': [item_code],
                'TOTAL_IN': [df_stock['DEBITQTY'].sum()],
                'TOTAL_OUT': [df_stock['CREDITQTY'].sum()],
                'CURRENT_STOCK': [df_stock['DEBITQTY'].sum() - df_stock['CREDITQTY'].sum()],
                'STOCK_TRANSACTIONS': [len(df_stock)],
                'SELECTED_SITES': [site_names_display]
            })
        else:
            # All sites - show by site
            result_df = df_stock.groupby('SITE').agg({
                'DEBITQTY': 'sum',
                'CREDITQTY': 'sum',
                'ITEM': 'first'
            }).reset_index()
            result_df['CURRENT_STOCK'] = result_df['DEBITQTY'] - result_df['CREDITQTY']
            result_df['STOCK_TRANSACTIONS'] = df_stock.groupby('SITE').size().values
            result_df = result_df.rename(columns={'DEBITQTY': 'TOTAL_IN', 'CREDITQTY': 'TOTAL_OUT'})
            result_df['SELECTED_SITES'] = ''
            result_df = result_df[['SITE', 'ITEM', 'TOTAL_IN', 'TOTAL_OUT', 'CURRENT_STOCK', 'STOCK_TRANSACTIONS', 'SELECTED_SITES']]
    elif site_code:
        # All items, single site
        result_df = df_stock.groupby('ITEM').agg({
            'DEBITQTY': 'sum',
            'CREDITQTY': 'sum',
            'SITE': 'first'
        }).reset_index()
        result_df['CURRENT_STOCK'] = result_df['DEBITQTY'] - result_df['CREDITQTY']
        result_df['STOCK_TRANSACTIONS'] = df_stock.groupby('ITEM').size().values
        result_df = result_df.rename(columns={'DEBITQTY': 'TOTAL_IN', 'CREDITQTY': 'TOTAL_OUT'})
        result_df['SELECTED_SITES'] = ''
        result_df = result_df[['SITE', 'ITEM', 'TOTAL_IN', 'TOTAL_OUT', 'CURRENT_STOCK', 'STOCK_TRANSACTIONS', 'SELECTED_SITES']]
    elif site_codes and isinstance(site_codes, list) and len(site_codes) > 0:
        # Multiple sites selected - aggregate by item across selected sites
        result_df = df_stock.groupby('ITEM').agg({
            'DEBITQTY': 'sum',
            'CREDITQTY': 'sum'
        }).reset_index()
        result_df['CURRENT_STOCK'] = result_df['DEBITQTY'] - result_df['CREDITQTY']
        result_df['STOCK_TRANSACTIONS'] = df_stock.groupby('ITEM').size().values
        result_df = result_df.rename(columns={'DEBITQTY': 'TOTAL_IN', 'CREDITQTY': 'TOTAL_OUT'})
        
        # Get site names for display
        site_names_display = "Multiple Sites"
        if 'sites' in dataframes and dataframes['sites'] is not None:
            site_name_map = dict(zip(dataframes['sites']['ID'], dataframes['sites']['SITE']))
            selected_site_names = [site_name_map.get(site_id, site_id) for site_id in site_codes]
            site_names_display = f"{', '.join(selected_site_names[:3])}" + (f" (+{len(selected_site_names)-3} more)" if len(selected_site_names) > 3 else "")
        
        # Add site columns
        result_df['SITE'] = f"Multiple Sites ({len(site_codes)})"
        result_df['SELECTED_SITES'] = site_names_display
        result_df = result_df[['SITE', 'ITEM', 'TOTAL_IN', 'TOTAL_OUT', 'CURRENT_STOCK', 'STOCK_TRANSACTIONS', 'SELECTED_SITES']]
    else:
        # All items, all sites
        result_df = df_stock.groupby(['SITE', 'ITEM']).agg({
            'DEBITQTY': 'sum',
            'CREDITQTY': 'sum'
        }).reset_index()
        result_df['CURRENT_STOCK'] = result_df['DEBITQTY'] - result_df['CREDITQTY']
        result_df['STOCK_TRANSACTIONS'] = df_stock.groupby(['SITE', 'ITEM']).size().values
        result_df = result_df.rename(columns={'DEBITQTY': 'TOTAL_IN', 'CREDITQTY': 'TOTAL_OUT'})
        result_df['SELECTED_SITES'] = ''
    
    # Limit results for performance in large datasets
    # Apply limit earlier and add sorting for better performance
    if len(result_df) > 1000:
        print(f"📊 Large dataset detected ({len(result_df)} rows), applying performance optimizations...")
        # Sort by current stock descending to get the most relevant items first
        result_df = result_df.sort_values('CURRENT_STOCK', ascending=False)
        result_df = result_df.head(1000)
        print(f"📊 Limited to top 1000 items by current stock for performance")
    
    print(f"📊 Processing {len(result_df)} rows for calculations...")
    
    # Calculate sales analytics from sales_details_df (ITEMS table)
    # Optimized version for better performance
    if 'sales_details' in dataframes and dataframes['sales_details'] is not None:
        df_sales = dataframes['sales_details']
        
        # Early filtering to reduce dataset size
        print("📊 Starting sales analytics calculation...")
        
        # Apply site filtering first to reduce data size
        sales_filters = []
        if item_code:
            sales_filters.append(df_sales['ITEM'] == item_code)
        if site_code:
            sales_filters.append(df_sales['SITE'] == site_code)
        if site_codes and isinstance(site_codes, list) and len(site_codes) > 0:
            sales_filters.append(df_sales['SITE'].isin(site_codes))
        
        # Apply all filters at once if any exist
        if sales_filters:
            combined_filter = sales_filters[0]
            for filt in sales_filters[1:]:
                combined_filter = combined_filter & filt
            df_sales = df_sales[combined_filter]
            print(f"📊 Filtered sales data to {len(df_sales)} records for analysis")
        
        # Continue with date and transaction type filtering
        if 'FDATE' in df_sales.columns and not df_sales.empty:
            df_sales = df_sales.copy()  # Only copy when needed
            df_sales['FDATE'] = pd.to_datetime(df_sales['FDATE'], errors='coerce')
            
            # Filter by date range if specified
            if from_date:
                from_date_dt = pd.to_datetime(from_date)
                df_sales = df_sales[df_sales['FDATE'] >= from_date_dt]
            if to_date:
                to_date_dt = pd.to_datetime(to_date)
                df_sales = df_sales[df_sales['FDATE'] <= to_date_dt]
          # Filter for sales transactions (FTYPE = 1 for sales, FTYPE = 2 for returns)
        if 'FTYPE' in df_sales.columns:
            sales_only = df_sales[df_sales['FTYPE'].isin([1, 2])]
        else:
            sales_only = df_sales
          # Filter for site sales only (SID starting with "530")
        if 'SID' in sales_only.columns:
            sales_only = sales_only[sales_only['SID'].astype(str).str.startswith('530')]
            print(f"📊 Filtered to {len(sales_only)} sales transactions with SID starting with '530' (site sales only)")
        else:
            print("⚠️ SID column not found - unable to filter for site sales")
        
        # Fill NaN values
        sales_only['QTY'] = sales_only['QTY'].fillna(0)
        
        # Calculate daily sales analytics
        def calculate_sales_analytics(group, from_date_param=None, to_date_param=None):
            if group.empty or 'FDATE' not in group.columns:
                return pd.Series({
                    'MAX_DAILY_SALES': 0,
                    'MIN_DAILY_SALES': 0, 
                    'AVG_DAILY_SALES': 0,
                    'SALES_TRANSACTIONS': 0,
                    'TOTAL_SALES_QTY': 0,
                    'SALES_PERIOD_DAYS': 0
                })
            
            # Group by date and sum quantities
            daily_sales = group.groupby('FDATE')['QTY'].sum()
            
            # Calculate period metrics
            total_sales_qty = group['QTY'].sum()
            
            # Use specified date range if provided, otherwise use actual sales date range
            if from_date_param and to_date_param:
                from_dt = pd.to_datetime(from_date_param)
                to_dt = pd.to_datetime(to_date_param)
                period_days = (to_dt - from_dt).days + 1
            elif from_date_param:
                from_dt = pd.to_datetime(from_date_param)
                max_date = group['FDATE'].max()
                period_days = (max_date - from_dt).days + 1
            elif to_date_param:
                min_date = group['FDATE'].min()
                to_dt = pd.to_datetime(to_date_param)
                period_days = (to_dt - min_date).days + 1
            else:
                # No date range specified, use actual sales period
                min_date = group['FDATE'].min()
                max_date = group['FDATE'].max()
                period_days = (max_date - min_date).days + 1 if min_date != max_date else 1
            
            # Exclude zero sales days for min calculation
            non_zero_sales = daily_sales[daily_sales > 0]
            
            max_sales = daily_sales.max() if not daily_sales.empty else 0
            min_sales = non_zero_sales.min() if not non_zero_sales.empty else 0
              # Calculate average daily sales using the full specified period (rounded to integer)
            avg_sales = round(total_sales_qty / period_days) if period_days > 0 else 0
            
            total_transactions = len(group)
            
            return pd.Series({
                'MAX_DAILY_SALES': max_sales,
                'MIN_DAILY_SALES': min_sales,
                'AVG_DAILY_SALES': avg_sales,
                'SALES_TRANSACTIONS': total_transactions,
                'TOTAL_SALES_QTY': total_sales_qty,
                'SALES_PERIOD_DAYS': period_days
            })
          # Calculate analytics by same grouping as stock
        # Optimized approach: skip sales analytics for multi-site aggregation to improve performance
        if site_codes and isinstance(site_codes, list) and len(site_codes) > 1:
            # Multi-site aggregation - skip detailed sales analytics to improve performance
            print("📊 Skipping detailed sales analytics for multi-site aggregation (performance optimization)")
            sales_analytics = pd.DataFrame()
        elif item_code and site_code:
            # Single item, single site
            if not sales_only.empty:
                analytics = calculate_sales_analytics(sales_only, from_date, to_date)
                sales_analytics = pd.DataFrame({
                    'SITE': [site_code],
                    'ITEM': [item_code],
                    'MAX_DAILY_SALES': [analytics['MAX_DAILY_SALES']],
                    'MIN_DAILY_SALES': [analytics['MIN_DAILY_SALES']], 
                    'AVG_DAILY_SALES': [analytics['AVG_DAILY_SALES']],
                    'SALES_TRANSACTIONS': [analytics['SALES_TRANSACTIONS']],
                    'TOTAL_SALES_QTY': [analytics['TOTAL_SALES_QTY']],
                    'SALES_PERIOD_DAYS': [analytics['SALES_PERIOD_DAYS']]
                })
            else:
                # No sales data - create empty DataFrame with correct structure
                sales_analytics = pd.DataFrame({
                    'SITE': [site_code],
                    'ITEM': [item_code],
                    'MAX_DAILY_SALES': [0],
                    'MIN_DAILY_SALES': [0], 
                    'AVG_DAILY_SALES': [0],
                    'SALES_TRANSACTIONS': [0],
                    'TOTAL_SALES_QTY': [0],
                    'SALES_PERIOD_DAYS': [0]
                })
        elif item_code:
            # Single item, all sites (but limit sites for performance)
            if not sales_only.empty:
                # Limit to top 20 sites by sales volume for performance
                top_sites = sales_only.groupby('SITE')['QTY'].sum().nlargest(20).index
                sales_only_limited = sales_only[sales_only['SITE'].isin(top_sites)]
                sales_analytics = sales_only_limited.groupby('SITE').apply(lambda x: calculate_sales_analytics(x, from_date, to_date), include_groups=False).reset_index()
                sales_analytics['ITEM'] = item_code
                print(f"📊 Limited sales analytics to top {len(top_sites)} sites for performance")
            else:
                # No sales data - create empty DataFrame
                sales_analytics = pd.DataFrame(columns=['SITE', 'ITEM', 'MAX_DAILY_SALES', 'MIN_DAILY_SALES', 'AVG_DAILY_SALES', 'SALES_TRANSACTIONS', 'TOTAL_SALES_QTY', 'SALES_PERIOD_DAYS'])
        elif site_code:
            # All items, single site (but limit items for performance)
            if not sales_only.empty:
                # Limit to top 100 items by sales volume for performance
                top_items = sales_only.groupby('ITEM')['QTY'].sum().nlargest(100).index
                sales_only_limited = sales_only[sales_only['ITEM'].isin(top_items)]
                sales_analytics = sales_only_limited.groupby('ITEM').apply(lambda x: calculate_sales_analytics(x, from_date, to_date), include_groups=False).reset_index()
                sales_analytics['SITE'] = site_code
                print(f"📊 Limited sales analytics to top {len(top_items)} items for performance")
            else:
                # No sales data - create empty DataFrame
                sales_analytics = pd.DataFrame(columns=['SITE', 'ITEM', 'MAX_DAILY_SALES', 'MIN_DAILY_SALES', 'AVG_DAILY_SALES', 'SALES_TRANSACTIONS', 'TOTAL_SALES_QTY', 'SALES_PERIOD_DAYS'])
        else:
            # All items, all sites - too expensive, skip for performance
            print("📊 Skipping detailed sales analytics for all items/all sites (performance optimization)")
            sales_analytics = pd.DataFrame()
          # Merge stock and sales analytics - only if we have sales data
        if not sales_analytics.empty and all(col in sales_analytics.columns for col in ['SITE', 'ITEM', 'MAX_DAILY_SALES', 'MIN_DAILY_SALES', 'AVG_DAILY_SALES', 'SALES_TRANSACTIONS', 'TOTAL_SALES_QTY', 'SALES_PERIOD_DAYS']):
            result_df = result_df.merge(sales_analytics[['SITE', 'ITEM', 'MAX_DAILY_SALES', 'MIN_DAILY_SALES', 'AVG_DAILY_SALES', 'SALES_TRANSACTIONS', 'TOTAL_SALES_QTY', 'SALES_PERIOD_DAYS']], 
                                       on=['SITE', 'ITEM'], how='left')
            result_df['MAX_DAILY_SALES'] = result_df['MAX_DAILY_SALES'].fillna(0)
            result_df['MIN_DAILY_SALES'] = result_df['MIN_DAILY_SALES'].fillna(0)
            result_df['AVG_DAILY_SALES'] = result_df['AVG_DAILY_SALES'].fillna(0)
            result_df['SALES_TRANSACTIONS'] = result_df['SALES_TRANSACTIONS'].fillna(0)
            result_df['TOTAL_SALES_QTY'] = result_df['TOTAL_SALES_QTY'].fillna(0)
            result_df['SALES_PERIOD_DAYS'] = result_df['SALES_PERIOD_DAYS'].fillna(0)
        else:
            # No sales analytics available - add empty sales columns
            print(f"⚠️ No sales analytics data to merge (sales_analytics empty: {sales_analytics.empty if 'sales_analytics' in locals() else 'undefined'})")
            result_df['MAX_DAILY_SALES'] = 0
            result_df['MIN_DAILY_SALES'] = 0
            result_df['AVG_DAILY_SALES'] = 0
            result_df['SALES_TRANSACTIONS'] = 0
            result_df['TOTAL_SALES_QTY'] = 0
            result_df['SALES_PERIOD_DAYS'] = 0
    else:
        # Add empty sales columns if sales data not available
        result_df['MAX_DAILY_SALES'] = 0
        result_df['MIN_DAILY_SALES'] = 0
        result_df['AVG_DAILY_SALES'] = 0
        result_df['SALES_TRANSACTIONS'] = 0
        result_df['TOTAL_SALES_QTY'] = 0
        result_df['SALES_PERIOD_DAYS'] = 0
    
    # Calculate stock autonomy (days of stock remaining at current sales rate)
    def calculate_autonomy(row):
        if row['CURRENT_STOCK'] <= 0:
            return -1  # Use -1 to represent N/A for zero stock
        elif row['AVG_DAILY_SALES'] > 0:
            return row['CURRENT_STOCK'] / row['AVG_DAILY_SALES']
        else:
            return 9999  # Use 9999 for infinite autonomy (stock but no sales)
    
    result_df['STOCK_AUTONOMY_DAYS'] = result_df.apply(calculate_autonomy, axis=1)
    
    # Calculate depot quantity (total stock available at depot sites where SIDNO = 3700004)
    try:
        if 'sites' not in dataframes or dataframes['sites'] is None:
            print("⚠️ Sites data not available for depot quantity calculation")
            result_df['DEPOT_QUANTITY'] = 0
        else:
            # Get depot sites where SIDNO = '3700004' (note: SIDNO is stored as string)
            depot_sites_info = dataframes['sites'][dataframes['sites']['SIDNO'] == '3700004'].copy()
            
            if depot_sites_info.empty:
                print("⚠️ No depot sites found with SIDNO = '3700004'")
                result_df['DEPOT_QUANTITY'] = 0
            else:
                depot_site_ids = depot_sites_info['ID'].tolist()
                print(f"📦 Found {len(depot_site_ids)} depot sites for quantity calculation")
                
                # Get inventory transactions for depot sites only
                df_depot = dataframes['inventory_transactions'][
                    dataframes['inventory_transactions']['SITE'].isin(depot_site_ids)
                ].copy()
                
                if df_depot.empty:
                    print("⚠️ No inventory transactions found for depot sites")
                    result_df['DEPOT_QUANTITY'] = 0
                else:
                    # Fill NaN values and calculate depot quantities by item
                    df_depot['DEBITQTY'] = df_depot['DEBITQTY'].fillna(0)
                    df_depot['CREDITQTY'] = df_depot['CREDITQTY'].fillna(0)
                    
                    # Calculate depot quantities by item (sum across all depot sites)
                    depot_qty = df_depot.groupby('ITEM').agg({
                        'DEBITQTY': 'sum',
                        'CREDITQTY': 'sum'
                    }).reset_index()
                    
                    depot_qty['DEPOT_QUANTITY'] = depot_qty['DEBITQTY'] - depot_qty['CREDITQTY']
                    
                    # Create dictionary for easy lookup
                    depot_dict = dict(zip(depot_qty['ITEM'], depot_qty['DEPOT_QUANTITY']))
                    
                    # Add depot quantity column to result_df
                    result_df['DEPOT_QUANTITY'] = result_df['ITEM'].map(depot_dict).fillna(0)
                    
                    total_depot_qty = depot_qty['DEPOT_QUANTITY'].sum()
                    items_with_depot_stock = (depot_qty['DEPOT_QUANTITY'] > 0).sum()
                    print(f"📊 Depot calculation: {total_depot_qty:,.0f} total units across {items_with_depot_stock:,} items")
                    
    except Exception as e:
        print(f"❌ Error calculating depot quantities: {e}")
        result_df['DEPOT_QUANTITY'] = 0
    
    # Exclude items with no sales in the last 6 months (except for depot sites)
    # Optimized version to improve performance
    six_months_ago = pd.to_datetime('today') - pd.DateOffset(months=6)
    
    # Check if we have sales data to filter
    if 'sales_details' in dataframes and dataframes['sales_details'] is not None:
        print("📊 Starting optimized 6-month filtering...")
        
        # Get the relevant sites first to reduce data size
        sites_to_check = []
        if site_codes and isinstance(site_codes, list) and len(site_codes) > 0:
            sites_to_check = site_codes
        elif site_code:
            sites_to_check = [site_code]
        
        # Identify depot sites early
        depot_items = set()
        is_depot_site_selected = False
        if 'sites' in dataframes and dataframes['sites'] is not None:
            depot_sites_info = dataframes['sites'][dataframes['sites']['SIDNO'] == '3700004'].copy()
            if not depot_sites_info.empty:
                depot_site_ids = depot_sites_info['ID'].tolist()
                
                # Check if any of our selected sites are depot sites
                if sites_to_check:
                    selected_depot_sites = [site for site in sites_to_check if site in depot_site_ids]
                    if selected_depot_sites:
                        is_depot_site_selected = True
                        depot_items = set(result_df['ITEM'].unique())
                        print(f"� Selected sites include depot sites - keeping all {len(depot_items)} items regardless of sales")
                else:
                    print("📦 No specific sites selected - applying 6-month sales filter")
        
        # Only do expensive sales filtering if depot sites are not selected
        if not is_depot_site_selected:
            # Build optimized query to minimize data processing
            df_sales = dataframes['sales_details']
            
            # Convert FDATE to datetime if it's not already (do this once)
            if 'FDATE' in df_sales.columns:
                print("📊 Converting dates and applying filters...")
                
                # Create boolean masks for efficient filtering
                date_mask = pd.to_datetime(df_sales['FDATE'], errors='coerce') >= six_months_ago
                
                # Filter for sales transactions (FTYPE = 1 for sales, FTYPE = 2 for returns)
                ftype_mask = df_sales['FTYPE'].isin([1, 2]) if 'FTYPE' in df_sales.columns else pd.Series([True] * len(df_sales))
                
                # Filter for site sales only (SID starting with "530")
                sid_mask = df_sales['SID'].astype(str).str.startswith('530') if 'SID' in df_sales.columns else pd.Series([True] * len(df_sales))
                
                # Filter by selected sites if specified
                if sites_to_check:
                    site_mask = df_sales['SITE'].isin(sites_to_check)
                else:
                    site_mask = pd.Series([True] * len(df_sales))
                
                # Combine all masks and apply at once
                combined_mask = date_mask & ftype_mask & sid_mask & site_mask
                
                print(f"📊 Applying combined filters to {len(df_sales)} total sales records...")
                df_sales_filtered = df_sales[combined_mask]
                
                print(f"📊 Filtered to {len(df_sales_filtered)} relevant sales transactions")
                
                # Get unique items that had sales in the last 6 months
                items_with_recent_sales = df_sales_filtered['ITEM'].unique()
                print(f"� Found {len(items_with_recent_sales)} items with recent sales")
                
                # Combine with depot items
                items_to_keep = set(items_with_recent_sales) | depot_items
            else:
                print("⚠️ FDATE column not found in sales data")
                items_to_keep = set(result_df['ITEM'].unique())
        else:
            # Depot site selected - keep all items
            items_to_keep = depot_items
        
        # Filter result_df to only include items with sales in the last 6 months OR depot sites
        result_df = result_df[result_df['ITEM'].isin(items_to_keep)]
        
        print(f"📊 Final result: {len(result_df)} items after 6-month filtering (since {six_months_ago.strftime('%Y-%m-%d')})")
    else:
        print("⚠️ No sales data available for 6-month filtering")
    
    # Add site names if available
    if 'sites' in dataframes and dataframes['sites'] is not None:
        site_names = dataframes['sites'][['ID', 'SITE']].drop_duplicates()
        site_names = site_names.rename(columns={'ID': 'SITE', 'SITE': 'SITE_NAME'})
        result_df = result_df.merge(site_names, on='SITE', how='left')
        
        # For multi-site aggregation, set SITE_NAME to match SELECTED_SITES if available
        multi_site_mask = result_df['SITE'].str.contains('Multiple Sites', na=False)
        result_df.loc[multi_site_mask, 'SITE_NAME'] = result_df.loc[multi_site_mask, 'SELECTED_SITES']
    else:
        # If no site data available, set SITE_NAME to SITE
        result_df['SITE_NAME'] = result_df['SITE']
      # Add item names and categories if available
    if 'inventory_items' in dataframes and dataframes['inventory_items'] is not None:
        item_info = dataframes['inventory_items'][['ITEM', 'DESCR1', 'CATEGORY', 'SUNIT', 'POSPRICE1']].drop_duplicates()
        item_info['ITEM_NAME'] = item_info['DESCR1'].fillna('').astype(str)
          # Add category descriptions if available
        if 'categories' in dataframes and dataframes['categories'] is not None:
            # Get category descriptions from DETDESCR table
            category_descriptions = dataframes['categories'][['ID', 'DESCR']].drop_duplicates()
            
            # Convert ID to string to match CATEGORY column type
            category_descriptions['ID'] = category_descriptions['ID'].astype(str)
            category_descriptions = category_descriptions.rename(columns={'ID': 'CATEGORY', 'DESCR': 'CATEGORY_NAME'})
            
            # Ensure CATEGORY column is string type for merge
            item_info['CATEGORY'] = item_info['CATEGORY'].astype(str)
            
            # Merge item info with category descriptions
            item_info = item_info.merge(category_descriptions, on='CATEGORY', how='left')
            item_info['CATEGORY_NAME'] = item_info['CATEGORY_NAME'].fillna('').astype(str)
        else:
            item_info['CATEGORY_NAME'] = ''
        
        # Merge with result_df
        result_df = result_df.merge(
            item_info[['ITEM', 'ITEM_NAME', 'CATEGORY', 'CATEGORY_NAME', 'SUNIT', 'POSPRICE1']], 
            on='ITEM', how='left'
        )
        
        # Fill missing values
        result_df['ITEM_NAME'] = result_df['ITEM_NAME'].fillna('')
        result_df['CATEGORY'] = result_df['CATEGORY'].fillna('')
        result_df['CATEGORY_NAME'] = result_df['CATEGORY_NAME'].fillna('')
        result_df['SUNIT'] = result_df['SUNIT'].fillna('')
        result_df['POSPRICE1'] = result_df['POSPRICE1'].fillna(0)
        
        # Calculate stock value (Price × Current Stock)
        result_df['STOCK_VALUE'] = result_df['POSPRICE1'] * result_df['CURRENT_STOCK']
    else:
        # If no item data available, create empty columns
        result_df['ITEM_NAME'] = ''
        result_df['CATEGORY'] = ''
        result_df['CATEGORY_NAME'] = ''
        result_df['SUNIT'] = ''
        result_df['POSPRICE1'] = 0
        result_df['STOCK_VALUE'] = 0
    
    return result_df

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/autonomy')
def autonomy():
    return render_template('autonomy.html')

@app.route('/favicon.ico')
def favicon():
    # Redirect to static favicon file
    from flask import redirect, url_for
    return redirect(url_for('static', filename='favicon.ico'))

@app.route('/custom-reports')
def custom_reports():
    return render_template('custom_reports.html')

@app.route('/stock-by-site')
def stock_by_site():
    return render_template('stock_by_site.html')

@app.route('/ciment-report')
def ciment_report():
    """Render the Ciment Report page"""
    return render_template('ciment_report.html')

@app.route('/api/load-dataframes', methods=['POST'])
def api_load_dataframes():
    global cache_loading
    try:
        print(f"📡 API load-dataframes called. Current cache_loading: {cache_loading}")
        
        with cache_lock:
            if cache_loading:
                print("⏳ Cache loading already in progress")
                return jsonify({'status': 'loading', 'message': 'Cache loading already in progress'})
            cache_loading = True
            print("🔒 Cache loading lock acquired")
        
        print("🚀 Starting dataframes loading...")
        result = load_dataframes()
        
        if result and len(result) > 0:
            print(f"✅ Successfully loaded {len(result)} dataframes")
            return jsonify({
                'status': 'success', 
                'message': f'Successfully loaded {len(result)} tables',
                'tables': list(result.keys())
            })
        else:
            print("❌ No dataframes were loaded")
            return jsonify({'status': 'error', 'message': 'No tables were loaded successfully'}), 500
            
    except Exception as e:
        cache_loading = False
        error_msg = str(e)
        print(f"❌ Error in api_load_dataframes: {error_msg}")
        return jsonify({'status': 'error', 'message': error_msg}), 500

@app.route('/api/cache-status')
def api_cache_status():
    return jsonify({
        'cached': len(dataframes) > 0,
        'loading': cache_loading,
        'tables': list(dataframes.keys()) if dataframes else []
    })

@app.route('/api/sites')
def api_sites():
    try:
        if 'sites' not in dataframes or dataframes['sites'] is None:
            return jsonify([])
        
        sites = dataframes['sites'][['ID', 'SITE']].drop_duplicates()
        sites_list = [{'id': row['ID'], 'name': row['SITE']} for _, row in sites.iterrows()]
        return jsonify(sites_list)
    except Exception as e:
        print(f"Error in api_sites: {e}")
        return jsonify([])

@app.route('/api/categories')
def api_categories():
    try:
        if 'categories' not in dataframes or dataframes['categories'] is None:
            return jsonify([])
        
        # Filter out categories with 0 items
        if 'inventory_items' in dataframes and dataframes['inventory_items'] is not None:
            # Get categories that have items
            categories_with_items = dataframes['inventory_items']['CATEGORY'].dropna().unique()
            
            # Filter categories to only include those with items
            categories = dataframes['categories'][['ID', 'DESCR']].drop_duplicates()
            categories = categories[categories['ID'].isin(categories_with_items)]
            
            categories_list = [{'id': row['ID'], 'name': row['DESCR'] or ''} for _, row in categories.iterrows()]
            
            # Sort by name for better UX
            categories_list.sort(key=lambda x: x['name'])
            
            print(f"📋 Returning {len(categories_list)} categories with items")
            return jsonify(categories_list)
        else:
            # Fallback if inventory_items is not available
            categories = dataframes['categories'][['ID', 'DESCR']].drop_duplicates()
            categories_list = [{'id': row['ID'], 'name': row['DESCR'] or ''} for _, row in categories.iterrows()]
            categories_list.sort(key=lambda x: x['name'])
            return jsonify(categories_list)
    except Exception as e:
        print(f"Error in api_categories: {e}")
        return jsonify([])

@app.route('/api/items')
def api_items():
    try:
        if 'inventory_items' not in dataframes or dataframes['inventory_items'] is None:
            return jsonify([])
        
        items = dataframes['inventory_items'][['ITEM', 'DESCR1']].drop_duplicates()
        items_list = [{'code': row['ITEM'], 'name': row['DESCR1'] or ''} for _, row in items.iterrows()]
        return jsonify(items_list)
    except Exception as e:
        print(f"Error in api_items: {e}")
        return jsonify([])

@app.route('/api/autonomy-report', methods=['POST'])
def api_autonomy_report():
    try:
        data = request.get_json()
        item_code = data.get('item_code')
        site_code = data.get('site_code') 
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        result_df = calculate_stock_and_sales(
            item_code=item_code, 
            site_code=site_code, 
            from_date=from_date, 
            to_date=to_date
        )
        
        if result_df is None or result_df.empty:
            return jsonify({'error': 'No data found for the specified criteria'}), 404
        
        # Calculate number of days in the selected period
        if from_date and to_date:
            from_dt = pd.to_datetime(from_date)
            to_dt = pd.to_datetime(to_date)
            period_days = (to_dt - from_dt).days + 1
        else:
            # Use the actual sales period from the data
            period_days = result_df['SALES_PERIOD_DAYS'].iloc[0] if not result_df.empty and 'SALES_PERIOD_DAYS' in result_df.columns else 0
        
        # Convert to JSON
        result_json = result_df.to_dict('records')
        
        # Return data with metadata
        return jsonify({
            'data': result_json,
            'metadata': {
                'period_days': int(period_days),
                'from_date': from_date,
                'to_date': to_date,
                'total_rows': len(result_df)
            }
        })
        
    except Exception as e:
        print(f"Error in autonomy report: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/stock-by-site-report', methods=['POST'])
def api_stock_by_site_report():
    try:
        data = request.get_json()
        item_code = data.get('item_code')
        site_code = data.get('site_code')  # Single site (legacy support)
        site_codes = data.get('site_codes')  # Multiple sites (new)
        category_id = data.get('category_id')  # Category filter
        as_of_date = data.get('as_of_date')  # Date filter parameter
        
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Use the existing calculate_stock_and_sales function with new parameters
        result_df = calculate_stock_and_sales(
            item_code=item_code, 
            site_code=site_code,
            site_codes=site_codes,
            category_id=category_id,
            as_of_date=as_of_date
        )
        
        if result_df is None or result_df.empty:
            return jsonify({'error': 'No data found for the specified criteria'}), 404
        
        # Sort by site name, then by current stock descending
        # For multi-site aggregation, sort by item name instead of site name
        if site_codes and isinstance(site_codes, list) and len(site_codes) > 1:
            # Multiple sites selected - sort by item name, then by current stock descending
            result_df = result_df.sort_values(['ITEM_NAME', 'CURRENT_STOCK'], ascending=[True, False])
        else:
            # Single site or all sites - sort by site name, then by current stock descending
            result_df = result_df.sort_values(['SITE_NAME', 'CURRENT_STOCK'], ascending=[True, False])
        
        # Convert to JSON
        result_json = result_df.to_dict('records')
        
        # Return data with metadata
        return jsonify({
            'data': result_json,
            'metadata': {
                'total_rows': len(result_df),
                'filters': {
                    'site_code': site_code,
                    'site_codes': site_codes,
                    'item_code': item_code,
                    'category_id': category_id,
                    'as_of_date': as_of_date
                }
            }
        })
        
    except Exception as e:
        print(f"Error in stock by site report: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-stock-by-site', methods=['POST'])
def api_export_stock_by_site():
    try:
        data = request.get_json()
        report_data = data.get('data', [])
        filters = data.get('filters', {})
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 400
        
        df = pd.DataFrame(report_data)
        
        # Build title with filters
        title = "Stock by Site Report"
        filter_info = []
        
        if filters.get('item_code'):
            filter_info.append(f"Item: {filters['item_code']}")
        
        # Handle site names display
        if filters.get('site_codes') and len(filters['site_codes']) > 0:
            site_codes = filters['site_codes']
            
            # Get site names from dataframes if available
            if 'sites' in dataframes and dataframes['sites'] is not None:
                site_info = dataframes['sites'][dataframes['sites']['ID'].isin(site_codes)]
                if not site_info.empty:
                    site_names = site_info['SITE'].tolist()
                    
                    if len(site_names) == 1:
                        filter_info.append(f"Site: {site_names[0]}")
                    elif len(site_names) <= 3:
                        filter_info.append(f"Sites: {', '.join(site_names)}")
                    else:
                        # Show first 3 sites and indicate there are more
                        displayed_sites = ', '.join(site_names[:3])
                        filter_info.append(f"Sites: {displayed_sites} (+{len(site_names)-3} more)")
                else:
                    # Fallback to site codes if names not found
                    if len(site_codes) == 1:
                        filter_info.append(f"Site: {site_codes[0]}")
                    else:
                        filter_info.append(f"Sites: {len(site_codes)} selected")
            else:
                # Fallback if sites dataframe not available
                if len(site_codes) == 1:
                    filter_info.append(f"Site: {site_codes[0]}")
                else:
                    filter_info.append(f"Sites: {len(site_codes)} selected")
        
        if filters.get('category_id'):
            filter_info.append(f"Category: {filters['category_id']}")
        if filters.get('as_of_date'):
            filter_info.append(f"As of Date: {filters['as_of_date']}")
        
        if filter_info:
            title += f" - {' | '.join(filter_info)}"
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock by Site', index=False, startrow=2)
            
            # Get workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['Stock by Site']
            
            # Add title with filters
            worksheet['A1'] = title
            worksheet['A1'].font = Font(size=14, bold=True)
            
            # Merge cells across all columns in the dataset
            # Use the DataFrame column count to determine the last column
            from openpyxl.utils import get_column_letter
            last_column = len(df.columns)  # Use DataFrame column count instead of worksheet.max_column
            last_column_letter = get_column_letter(last_column)
            
            if last_column > 1:
                worksheet.merge_cells(f'A1:{last_column_letter}1')
                print(f"📊 Merged cells A1:{last_column_letter}1 for {last_column} columns")
            else:
                worksheet.merge_cells('A1:A1')
            
            # Center the title text
            worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Style the header row (now at row 3)
            header_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create descriptive filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['stock_by_site_report']
        
        # Add sites to filename if specified
        if filters.get('site_codes') and len(filters['site_codes']) > 0:
            filename_parts.append(f'sites_{len(filters["site_codes"])}')
        
        # Add category to filename if specified
        if filters.get('category_id'):
            category_id = str(filters['category_id']).replace(' ', '_').replace('/', '_')
            filename_parts.append(f'category_{category_id}')
        
        # Add date to filename if specified
        if filters.get('as_of_date'):
            as_of_date_str = filters['as_of_date'].replace('-', '')
            filename_parts.append(f'as_of_{as_of_date_str}')
        
        # Add timestamp
        filename_parts.append(timestamp)
        
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
def api_export_excel():
    try:
        data = request.get_json()
        report_data = data.get('data', [])
        filters = data.get('filters', {})
        
        # Debug logging
        print(f"📊 Export Excel - Received filters: {filters}")
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 400
        
        df = pd.DataFrame(report_data)
        
        # Get site name if site_code is specified
        site_name = None
        if filters.get('site_code') and 'sites' in dataframes and dataframes['sites'] is not None:
            site_info = dataframes['sites'][dataframes['sites']['ID'] == filters['site_code']]
            if not site_info.empty:
                site_name = site_info['SITE'].iloc[0]
                print(f"📊 Found site name: {site_name} for site code: {filters['site_code']}")
        
        # Build title with filters
        title = "Stock Autonomy Report"
        filter_info = []
        
        if filters.get('item_code'):
            filter_info.append(f"Item: {filters['item_code']}")
        if filters.get('site_code'):
            if site_name:
                filter_info.append(f"Site: {site_name}")
            else:
                filter_info.append(f"Site: {filters['site_code']}")
        if filters.get('from_date'):
            filter_info.append(f"From: {filters['from_date']}")
        if filters.get('to_date'):
            filter_info.append(f"To: {filters['to_date']}")
        
        if filter_info:
            title += f" - {' | '.join(filter_info)}"
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock Autonomy', index=False, startrow=2)
            
            # Get workbook and worksheet for styling
            workbook = writer.book
            worksheet = writer.sheets['Stock Autonomy']
            
            # Add title with filters
            worksheet['A1'] = title
            worksheet['A1'].font = Font(size=14, bold=True)
            
            # Merge cells across all columns in the dataset
            # Use the DataFrame column count to determine the last column
            from openpyxl.utils import get_column_letter
            last_column = len(df.columns)  # Use DataFrame column count instead of worksheet.max_column
            last_column_letter = get_column_letter(last_column)
            
            if last_column > 1:
                worksheet.merge_cells(f'A1:{last_column_letter}1')
                print(f"📊 Merged cells A1:{last_column_letter}1 for {last_column} columns")
            else:
                worksheet.merge_cells('A1:A1')
            
            # Center the title text
            worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Style the header row (now at row 3)
            header_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=3, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Create descriptive filename with date and site information
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['stock_autonomy_report']
        
        # Debug logging
        print(f"📊 Building filename - filters: {filters}")
        
        # Add site (with full name) to filename if specified
        if filters.get('site_code'):
            if site_name:
                # Clean site name for filename (remove special characters)
                clean_site_name = site_name.replace(' ', '_').replace('/', '_').replace('\\', '_').replace(':', '_').replace('?', '').replace('*', '').replace('<', '').replace('>', '').replace('|', '').replace('"', '')
                filename_parts.append(f'site_{clean_site_name}')
                print(f"📊 Added site to filename: site_{clean_site_name}")
            else:
                site_code = filters['site_code'].replace(' ', '_').replace('/', '_')
                filename_parts.append(f'site_{site_code}')
                print(f"📊 Added site code to filename: site_{site_code}")
        
        # Add date range to filename if specified
        if filters.get('from_date') and filters.get('to_date'):
            from_date_str = filters['from_date'].replace('-', '')
            to_date_str = filters['to_date'].replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
            print(f"📊 Added date range to filename: {from_date_str}_to_{to_date_str}")
        elif filters.get('from_date'):
            from_date_str = filters['from_date'].replace('-', '')
            filename_parts.append(f'from_{from_date_str}')
            print(f"📊 Added from date to filename: from_{from_date_str}")
        elif filters.get('to_date'):
            to_date_str = filters['to_date'].replace('-', '')
            filename_parts.append(f'to_{to_date_str}')
            print(f"📊 Added to date to filename: to_{to_date_str}")
        
        # Add timestamp
        filename_parts.append(timestamp)
        
        filename = '_'.join(filename_parts) + '.xlsx'
        print(f"📊 Final filename: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get-available-columns')
def api_get_available_columns():
    """Get all available columns from all loaded dataframes"""
    try:
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        available_tables = {}
        
        # Table descriptions for better user experience
        table_descriptions = {
            'sites': 'Site/Location Master Data',
            'categories': 'Category Definitions', 
            'invoice_headers': 'Invoice Headers',
            'sales_details': 'Sales Transaction Details',
            'vouchers': 'Payment Vouchers',
            'accounts': 'Statement of Accounts',
            'inventory_items': 'Items/Products Master',
            'inventory_transactions': 'All Inventory Transactions'
        }
        
        for table_name, df in dataframes.items():
            if df is not None and not df.empty:
                columns_list = []
                for col in df.columns:
                    # Get sample data type
                    dtype = str(df[col].dtype)
                    
                    # Get sample non-null value for context
                    sample_value = None
                    non_null_values = df[col].dropna()
                    if not non_null_values.empty:
                        sample_value = str(non_null_values.iloc[0])[:50]  # Limit length
                    
                    columns_list.append(col)
                
                available_tables[table_name] = {
                    'description': table_descriptions.get(table_name, table_name.replace('_', ' ').title()),
                    'columns': columns_list,
                    'row_count': len(df)
                }
        
        return jsonify(available_tables)
        
    except Exception as e:
        print(f"Error in get_available_columns: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/custom-report', methods=['POST'])
def api_custom_report():
    try:
        data = request.get_json()
        selected_tables = data.get('tables', [])  # Which tables to include
        selected_columns = data.get('columns', [])
        calculated_fields = data.get('calculated_fields', [])
        filters = data.get('filters', [])
        sort_config = data.get('sort', {})
        join_config = data.get('joins', [])  # How to join tables
        
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Check if we need calculated columns (stock and sales analysis)
        needs_calculation = False
        calc_columns = {
            'MAX_DAILY_SALES', 'MIN_DAILY_SALES', 'AVG_DAILY_SALES', 
            'SALES_TRANSACTIONS', 'TOTAL_SALES_QTY', 'SALES_PERIOD_DAYS',
            'CURRENT_STOCK', 'STOCK_AUTONOMY_DAYS', 'DEPOT_QUANTITY',
            'TOTAL_IN', 'TOTAL_OUT', 'STOCK_TRANSACTIONS'
        }
        
        for col in selected_columns:
            column_name = col.split('.')[-1] if '.' in col else col
            if column_name in calc_columns:
                needs_calculation = True
                break
        
        # If no tables specified or calculated columns needed, use the stock and sales calculation
        if not selected_tables or needs_calculation:
            result_df = calculate_stock_and_sales()
            if result_df is None or result_df.empty:
                return jsonify({'error': 'No data found'}), 404
        else:
            # Build custom dataset from selected tables
            result_df = None
            
            # Parse selected columns to determine which tables we need
            tables_needed = set()
            column_mapping = {}
            
            for col in selected_columns:
                if '.' in col:
                    table_name, column_name = col.split('.', 1)
                    tables_needed.add(table_name)
                    column_mapping[col] = (table_name, column_name)
            
            # Load and join required tables
            for table_name in tables_needed:
                if table_name not in dataframes or dataframes[table_name] is None:
                    continue
                
                # Get the table data
                table_df = dataframes[table_name].copy()
                
                # Limit rows for performance
                table_df = table_df.head(10000)
                
                if result_df is None:
                    result_df = table_df
                    # Rename columns to include table prefix
                    result_df.columns = [f"{table_name}.{col}" for col in result_df.columns]
                else:
                    # Try to find a common key for joining
                    common_keys = []
                    for col in result_df.columns:
                        base_col = col.split('.')[-1]  # Get column name without prefix
                        if base_col in table_df.columns:
                            common_keys.append((col, base_col))
                    
                    if common_keys:
                        # Use the first common key found
                        left_key, right_key = common_keys[0]
                        
                        # Rename columns in new table to include prefix
                        table_df_renamed = table_df.copy()
                        table_df_renamed.columns = [f"{table_name}.{col}" for col in table_df_renamed.columns]
                        right_key_renamed = f"{table_name}.{right_key}"
                        
                        result_df = result_df.merge(
                            table_df_renamed,
                            left_on=left_key,
                            right_on=right_key_renamed,
                            how='left'
                        )
                    else:
                        print(f"Warning: No common key found for joining {table_name}")
            
            if result_df is None or result_df.empty:
                return jsonify({'error': 'No data found for selected tables'}), 404
          # Apply filters
        for filter_config in filters:
            column = filter_config.get('column')
            operator = filter_config.get('operator')
            value = filter_config.get('value')
            
            if column in result_df.columns:
                try:
                    if operator == 'equals':
                        result_df = result_df[result_df[column] == value]
                    elif operator == 'contains':
                        result_df = result_df[result_df[column].astype(str).str.contains(str(value), case=False, na=False)]
                    elif operator == 'greater_than':
                        result_df = result_df[pd.to_numeric(result_df[column], errors='coerce') > float(value)]
                    elif operator == 'less_than':
                        result_df = result_df[pd.to_numeric(result_df[column], errors='coerce') < float(value)]
                    elif operator == 'between':
                        min_val, max_val = value.split(',')
                        result_df = result_df[
                            (pd.to_numeric(result_df[column], errors='coerce') >= float(min_val)) &
                            (pd.to_numeric(result_df[column], errors='coerce') <= float(max_val))
                        ]
                except Exception as e:
                    print(f"Warning: Error applying filter on column {column}: {e}")
        
        # Add calculated fields
        for calc_field in calculated_fields:
            field_name = calc_field.get('name')
            formula = calc_field.get('formula')
            
            try:
                # Simple formula evaluation (extend this for more complex formulas)
                # For now, support basic arithmetic operations with column references
                eval_formula = formula
                for col in result_df.columns:
                    # Replace both formats: [column] and [table.column]
                    eval_formula = eval_formula.replace(f'[{col}]', f'result_df["{col}"]')
                    if '.' in col:
                        short_col = col.split('.')[-1]
                        eval_formula = eval_formula.replace(f'[{short_col}]', f'result_df["{col}"]')
                
                # Use eval with restricted globals for security
                result_df[field_name] = eval(eval_formula, {'__builtins__': {}, 'pd': pd, 'np': np}, {'result_df': result_df})
            except Exception as e:
                print(f"Warning: Error in calculated field '{field_name}': {e}")
          # Select only requested columns
        if selected_columns:
            final_columns = []
            
            for col in selected_columns:
                if needs_calculation:
                    # For calculated data, columns don't have table prefixes
                    column_name = col.split('.')[-1] if '.' in col else col
                    if column_name in result_df.columns:
                        final_columns.append(column_name)
                else:
                    # For raw table data, columns have table prefixes
                    if col in result_df.columns:
                        final_columns.append(col)
                    else:
                        # Try to find column with prefix
                        column_name = col.split('.')[-1] if '.' in col else col
                        matching_cols = [c for c in result_df.columns if c.endswith(f".{column_name}")]
                        if matching_cols:
                            final_columns.append(matching_cols[0])
                        else:
                            # Last resort: try without any prefix
                            if column_name in result_df.columns:
                                final_columns.append(column_name)
            
            if final_columns:
                result_df = result_df[final_columns]
                
                # Rename columns to remove table prefixes for display
                display_columns = {}
                for col in final_columns:
                    if '.' in col:
                        table_name, column_name = col.split('.', 1)
                        display_columns[col] = f"{column_name} ({table_name})"
                    else:
                        display_columns[col] = col
                
                result_df = result_df.rename(columns=display_columns)
            else:
                return jsonify({'error': 'No matching columns found in the data'}), 400
          # Apply sorting
        if sort_config.get('column'):
            sort_column = sort_config['column']
            # The sort column should match the display column names now
            actual_sort_column = None
            
            # Look for exact match first
            if sort_column in result_df.columns:
                actual_sort_column = sort_column
            else:
                # Try to find by original column name pattern
                for col in result_df.columns:
                    # Check if it's a renamed column that matches our sort target
                    if col.startswith(sort_column.split('.')[-1]):
                        actual_sort_column = col
                        break
                    # Or check if the sort column matches the display name pattern
                    if sort_column in col:
                        actual_sort_column = col
                        break
            
            if actual_sort_column:
                ascending = sort_config.get('direction', 'asc') == 'asc'
                result_df = result_df.sort_values(actual_sort_column, ascending=ascending)
        
        # Convert to JSON
        result_json = result_df.to_dict('records')
        
        return jsonify({
            'data': result_json,
            'columns': list(result_df.columns),
            'metadata': {
                'total_rows': len(result_df),
                'selected_columns': len(selected_columns) if selected_columns else len(result_df.columns),
                'calculated_fields': len(calculated_fields),
                'filters_applied': len(filters)
            }        })
        
    except Exception as e:
        print(f"Error in custom report: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/custom-report-export', methods=['POST'])
def api_custom_report_export():
    try:
        data = request.get_json()
        export_format = data.get('format', 'csv').lower()
        report_data = data.get('report_data', [])
        columns = data.get('columns', [])
        report_title = data.get('title', 'Custom Report')
        filters = data.get('filters', [])
        calculated_fields = data.get('calculated_fields', [])
        
        if not report_data:
            return jsonify({'error': 'No report data to export'}), 400
        
        # Build title with filters and calculated fields
        title_parts = [report_title]
        
        # Add filter information
        if filters:
            filter_info = []
            for f in filters:
                operator_text = {
                    'equals': '=',
                    'contains': 'contains',
                    'greater_than': '>',
                    'less_than': '<',
                    'between': 'between'
                }.get(f['operator'], f['operator'])
                filter_info.append(f"{f['column']} {operator_text} {f['value']}")
            title_parts.append(f"Filters: {' | '.join(filter_info)}")
        
        # Add calculated fields information
        if calculated_fields:
            calc_info = [f"{cf['name']}: {cf['formula']}" for cf in calculated_fields]
            title_parts.append(f"Formulas: {' | '.join(calc_info)}")
        
        full_title = ' - '.join(title_parts)
        
        # Convert to DataFrame
        df = pd.DataFrame(report_data)
        
        if export_format == 'csv':
            # CSV Export with title
            output = io.StringIO()
            
            # Add title as first line
            output.write(f"# {full_title}\n")
            output.write(f"# Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            output.write("\n")
            
            # Add the actual data
            df.to_csv(output, index=False)
            output.seek(0)
            
            return send_file(
                io.BytesIO(output.getvalue().encode()),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'{report_title.replace(" ", "_")}.csv'

            )
            
        elif export_format == 'excel':
            # Excel Export with title and filters
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Report', index=False, startrow=3)
                
                # Get workbook and worksheet for styling
                workbook = writer.book
                worksheet = writer.sheets['Report']
                
                # Add title with filters
                worksheet['A1'] = full_title
                worksheet['A1'].font = Font(size=14, bold=True)
                worksheet.merge_cells('A1:D1')
                
                # Add generation timestamp
                worksheet['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                worksheet['A2'].font = Font(size=10, italic=True)
                worksheet.merge_cells('A2:D2')
                
                # Style the header row (now at row 4)
                header_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=4, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'{report_title.replace(" ", "_")}.xlsx'
            )
            
        elif export_format == 'pdf':
            # PDF Export with title and filters
            output = io.BytesIO()
            
            doc = SimpleDocTemplate(output, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                textColor=colors.HexColor('#1f4e79')
            )
            elements.append(Paragraph(full_title, title_style))
            
            # Generation timestamp
            timestamp_style = ParagraphStyle(
                'Timestamp',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=20,
                textColor=colors.grey
            )
            elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", timestamp_style))
            elements.append(Spacer(1, 12))
            
            # Create table data
            table_data = [list(df.columns)]  # Header
            for _, row in df.iterrows():
                table_data.append([str(val) for val in row.values])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'{report_title.replace(" ", "_")}.pdf'
            )
            
        else:
            return jsonify({'error': 'Unsupported export format'}), 400
            
    except Exception as e:
        print(f"Error in export: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/ciment-report', methods=['POST'])
def api_ciment_report():
    """
    Generate Ciment Report showing site-wise ciment category sales and stock
    """
    try:
        data = request.get_json()
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get all non-depot sites (exclude depot sites where SIDNO = '3700004')
        if 'sites' not in dataframes or dataframes['sites'] is None:
            return jsonify({'error': 'Sites data not available'}), 400
        
        # Get Kinshasa sites only (SIDNO = '3700002')
        kinshasa_sites = dataframes['sites'][dataframes['sites']['SIDNO'] == '3700002'].copy()
        
        if kinshasa_sites.empty:
            return jsonify({'error': 'No Kinshasa sites found (SIDNO = 3700002)'}), 404
            
        print(f"📍 Found {len(kinshasa_sites)} Kinshasa sites for ciment report")
        
        # Get ciment category items
        if 'categories' not in dataframes or dataframes['categories'] is None:
            return jsonify({'error': 'Categories data not available'}), 400
        
        # Find ciment category (look for category containing "ciment" in name)
        ciment_categories = dataframes['categories'][
            dataframes['categories']['DESCR'].str.contains('ciment', case=False, na=False)
        ]
        
        if ciment_categories.empty:
            # Try alternative names
            ciment_categories = dataframes['categories'][
                dataframes['categories']['DESCR'].str.contains('cement', case=False, na=False)
            ]
        
        if ciment_categories.empty:
            return jsonify({'error': 'Ciment category not found. Please check category names.'}), 404
        
        ciment_category_id = ciment_categories.iloc[0]['ID']
        print(f"📋 Found ciment category: {ciment_categories.iloc[0]['DESCR']} (ID: {ciment_category_id})")
        
        # Get items in ciment category
        if 'inventory_items' not in dataframes or dataframes['inventory_items'] is None:
            return jsonify({'error': 'Inventory items data not available'}), 400
        
        # Get all ciment items from the ciment category
        ciment_items = dataframes['inventory_items'][
            dataframes['inventory_items']['CATEGORY'].astype(str) == str(ciment_category_id)
        ]['ITEM'].unique()
        
        if len(ciment_items) == 0:
            return jsonify({'error': 'No items found in ciment category'}), 404
        
        print(f"📦 Found {len(ciment_items)} items in ciment category")
        
        # Pre-filter sales data to find items with actual sales
        if 'sales_details' in dataframes and dataframes['sales_details'] is not None:
            all_sales = dataframes['sales_details'].copy()
            
            # Filter by date range if specified
            if from_date or to_date:
                all_sales['FDATE'] = pd.to_datetime(all_sales['FDATE'], errors='coerce')
                if from_date:
                    from_date_dt = pd.to_datetime(from_date)
                    all_sales = all_sales[all_sales['FDATE'] >= from_date_dt]
                if to_date:
                    to_date_dt = pd.to_datetime(to_date)
                    all_sales = all_sales[all_sales['FDATE'] <= to_date_dt]
            
            # Filter for sales transactions only (FTYPE = 1 for sales)
            if 'FTYPE' in all_sales.columns:
                all_sales = all_sales[all_sales['FTYPE'] == 1]
            
            # Filter for site sales only (SID starting with "530")
            if 'SID' in all_sales.columns:
                all_sales = all_sales[all_sales['SID'].astype(str).str.startswith('530')]
            
            # Find ciment items that have sales in the period
            ciment_items_with_sales = all_sales[
                all_sales['ITEM'].isin(ciment_items)
            ]['ITEM'].unique()
            
            # Use only ciment items that have sales
            active_ciment_items = ciment_items_with_sales
            
            print(f"📊 Pre-filtered sales data: {len(all_sales)} transactions")
            print(f"📦 Active ciment items (with sales): {len(active_ciment_items)} items shown")
        else:
            all_sales = pd.DataFrame()
            active_ciment_items = ciment_items
            print("⚠️ No sales data available - showing all ciment items")
        # Get item prices for amount calculations
        item_prices = {}
        if 'inventory_items' in dataframes and dataframes['inventory_items'] is not None:
            price_data = dataframes['inventory_items'][
                dataframes['inventory_items']['ITEM'].isin(active_ciment_items)
            ][['ITEM', 'POSPRICE1']].drop_duplicates()
            item_prices = dict(zip(price_data['ITEM'], price_data['POSPRICE1'].fillna(0)))
            print(f"📊 Retrieved prices for {len(item_prices)} ciment items")
        
        # Calculate sales for each site (no stock data needed)
        result_data = []
        
        for i, (_, site) in enumerate(kinshasa_sites.iterrows()):
            if i % 10 == 0:  # Progress logging every 10 sites
                print(f"📍 Processing site {i+1}/{len(kinshasa_sites)}: {site['SITE']}")
            
            site_id = site['ID']
            site_name = site['SITE']
            
            # Initialize row data with Python native types (sales only)
            row_data = {
                'SITE_ID': str(site_id),
                'SITE_NAME': str(site_name),
                'TOTAL_AMOUNT': 0.0,  # Total sales amount (qty × price)
                'CIMENT_ARTICLES_WITH_SALES': 0,  # Count of ciment items with sales
                'TOTAL_CIMENT_SALES_QTY': 0.0  # Total ciment sales quantity
            }
            
            # Calculate sales for this site
            if not all_sales.empty:
                site_sales = all_sales[all_sales['SITE'] == site_id]
                
                # Calculate ciment sales by item with amounts
                ciment_sales_by_item = {}
                ciment_amounts_by_item = {}
                total_ciment_sales = 0.0
                total_amount = 0.0
                articles_with_sales = 0
                
                for item_code in active_ciment_items:
                    item_sales = site_sales[site_sales['ITEM'] == item_code]
                    item_sales_qty = float(item_sales['QTY'].fillna(0).sum())
                    
                    # Calculate amount (qty × price)
                    item_price = item_prices.get(item_code, 0)
                    item_amount = item_sales_qty * item_price
                    
                    if item_sales_qty > 0:
                        articles_with_sales += 1
                    
                    ciment_sales_by_item[str(item_code)] = item_sales_qty
                    ciment_amounts_by_item[str(item_code)] = item_amount
                    total_ciment_sales += item_sales_qty
                    total_amount += item_amount
                
                row_data['TOTAL_AMOUNT'] = total_amount
                row_data['CIMENT_ARTICLES_WITH_SALES'] = articles_with_sales
                row_data['TOTAL_CIMENT_SALES_QTY'] = total_ciment_sales
                row_data['CIMENT_SALES_BY_ITEM'] = ciment_sales_by_item
                row_data['CIMENT_AMOUNTS_BY_ITEM'] = ciment_amounts_by_item
            else:
                row_data['CIMENT_SALES_BY_ITEM'] = {str(item): 0.0 for item in active_ciment_items}
                row_data['CIMENT_AMOUNTS_BY_ITEM'] = {str(item): 0.0 for item in active_ciment_items}
            
            result_data.append(row_data)
        
        # Sort by site name
        result_data.sort(key=lambda x: x['SITE_NAME'])
        
        # Get ciment item details for column headers (only active items)
        ciment_item_details = dataframes['inventory_items'][
            dataframes['inventory_items']['ITEM'].isin(active_ciment_items)
        ][['ITEM', 'DESCR1']].drop_duplicates()
        
        # Convert to JSON-serializable format
        ciment_item_details = [
            {
                'ITEM': str(row['ITEM']), 
                'DESCR1': str(row['DESCR1']) if pd.notna(row['DESCR1']) else ''
            } 
            for _, row in ciment_item_details.iterrows()
        ]
        
        return jsonify({
            'data': result_data,
            'metadata': {
                'total_sites': len(result_data),
                'ciment_category_id': int(ciment_category_id),  # Convert to Python int
                'ciment_category_name': str(ciment_categories.iloc[0]['DESCR']),  # Convert to Python string
                'ciment_items': ciment_item_details,
                'total_ciment_items_in_category': len(ciment_items),
                'active_ciment_items_shown': len(active_ciment_items),
                'from_date': from_date,
                'to_date': to_date,
                'sales_only_mode': True,  # Indicate this is sales-only report
                'columns_info': {
                    'site': 'Site name',
                    'total_amount': 'Total amount (sales × price)',
                    'ciment_articles_with_sales': 'Number of ciment articles with sales',
                    'total_ciment_sales_qty': 'Total ciment sales quantity'
                }
            }
        })
        
    except Exception as e:
        print(f"Error in ciment report: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-ciment-report', methods=['POST'])
def api_export_ciment_report():
    """Export ciment report to Excel with proper formatting"""
    try:
        data = request.get_json()
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the ciment report data using the same logic as the API
        # Call the existing ciment report API logic
        from flask import current_app
        with current_app.test_request_context('/api/ciment-report', method='POST', json=data):
            response = api_ciment_report()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ciment Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        title = "Ciment Report"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        elif from_date:
            title += f" - From: {from_date}"
        elif to_date:
            title += f" - To: {to_date}"
        
        # Write title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Total Sites: {len(report_data)}"
        row += 1
        ws[f'A{row}'] = f"Ciment Category: All Ciment Items (Sales Only - Dynamic)"
        row += 1
        if metadata.get('active_ciment_items_shown') and metadata.get('total_ciment_items_in_category'):
            active_count = metadata.get('active_ciment_items_shown')
            total_count = metadata.get('total_ciment_items_in_category')
            ws[f'A{row}'] = f"Active Ciment Items: {active_count} of {total_count} ciment items with sales data"
        else:
            ws[f'A{row}'] = f"Ciment Items: {len(metadata.get('ciment_items', []))}"
        row += 1
        if metadata.get('sales_only_mode'):
            ws[f'A{row}'] = "Note: This report shows sales data only with amounts (qty × price)"
        row += 2
        
        # Headers (new structure: site, amount, articles with sales, total qty, individual items)
        headers = [
            'Site Name',
            'Total Amount',
            'Articles with Sales',
            'Total Sales Qty'
        ]
        
        # Add individual ciment item sales columns only
        if metadata.get('ciment_items'):
            for item in metadata['ciment_items']:
                # Create compact header: "ITEM_CODE - ITEM_NAME" (truncate name if too long)
                item_code = item['ITEM']
                item_name = item['DESCR1']
                
                # Truncate item name if it's too long to save width
                if len(item_name) > 25:
                    item_name = item_name[:22] + "..."
                
                compact_header = f"{item_code} - {item_name}" if item_name else item_code
                headers.append(compact_header)
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        row += 1
        
        # Write data (new structure with amounts and article counts)
        for site_data in report_data:
            col = 1
            
            # Basic columns (new structure)
            ws.cell(row=row, column=col, value=site_data.get('SITE_NAME', '')).border = border
            col += 1
            ws.cell(row=row, column=col, value=site_data.get('TOTAL_AMOUNT', 0)).border = border
            col += 1
            ws.cell(row=row, column=col, value=site_data.get('CIMENT_ARTICLES_WITH_SALES', 0)).border = border
            col += 1
            ws.cell(row=row, column=col, value=site_data.get('TOTAL_CIMENT_SALES_QTY', 0)).border = border
            col += 1
            
            # Individual item sales only (no stock)
            if metadata.get('ciment_items'):
                for item in metadata['ciment_items']:
                    sales_value = 0
                    if site_data.get('CIMENT_SALES_BY_ITEM') and item['ITEM'] in site_data['CIMENT_SALES_BY_ITEM']:
                        sales_value = site_data['CIMENT_SALES_BY_ITEM'][item['ITEM']]
                    ws.cell(row=row, column=col, value=sales_value).border = border
                    col += 1
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row_num in range(1, ws.max_row + 1):
                cell_value = str(ws[f'{column_letter}{row_num}'].value or '')
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create filename
        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        filename = f'ciment_report_{timestamp}.xlsx'
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in ciment report export: {e}")
        return jsonify({'error': str(e)}), 500
    

if __name__ == '__main__':
    print("🚀 Starting DustReports Flask Application")
    print("📊 Using notebook-style data loading approach")
    app.run(host='0.0.0.0', port=5000, debug=True)
