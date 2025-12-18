"""Export routes for file downloads"""

from flask import Blueprint, request, jsonify, send_file
import pandas as pd
from utils.export_utils import ExcelExporter, PDFExporter
from services.database_service import get_dataframes

export_bp = Blueprint('export', __name__, url_prefix='/api')

@export_bp.route('/export-excel', methods=['POST'])
def api_export_excel():
    """Export autonomy report to Excel"""
    try:
        data = request.get_json()
        report_data = data.get('data', [])
        filters = data.get('filters', {})
        
        output, filename = ExcelExporter.create_stock_autonomy_report(report_data, filters)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-stock-by-site', methods=['POST'])
def api_export_stock_by_site():
    """Export stock by site report to Excel"""
    try:
        data = request.get_json()
        report_data = data.get('data', [])
        filters = data.get('filters', {})
        
        output, filename = ExcelExporter.create_stock_by_site_report(report_data, filters)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-pdf', methods=['POST'])
def api_export_pdf():
    """Export any report to PDF"""
    try:
        data = request.get_json()
        report_data = data.get('data', [])
        title = data.get('title', 'Report')
        filters = data.get('filters', {})
        
        output = PDFExporter.create_report(report_data, title, filters)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{title.lower().replace(' ', '_')}_{timestamp}.pdf"
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-ciment-report', methods=['POST'])
def api_export_ciment_report():
    """Export ciment report to Excel with proper formatting"""
    try:
        data = request.get_json()
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        
        dataframes = get_dataframes()
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the ciment report data by calling the API function directly
        from routes.api_routes import api_ciment_report
        from flask import current_app
        
        with current_app.test_request_context('/api/ciment-report', method='POST', json=data):
            response = api_ciment_report()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ciment Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        title = "Ciment Report"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        elif from_date:
            title += f" - From: {from_date}"
        elif to_date:
            title += f" - To: {to_date}"
        
        # Write title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Total Sites: {len(report_data)}"
        row += 1
        ws[f'A{row}'] = f"Ciment Category: All Ciment Items (Sales Only - Dynamic)"
        row += 1
        if metadata.get('active_ciment_items_shown') and metadata.get('total_ciment_items_in_category'):
            active_count = metadata.get('active_ciment_items_shown')
            total_count = metadata.get('total_ciment_items_in_category')
            ws[f'A{row}'] = f"Active Ciment Items: {active_count} of {total_count} ciment items with sales data"
        else:
            ws[f'A{row}'] = f"Ciment Items: {len(metadata.get('ciment_items', []))}"
        row += 1
        if metadata.get('sales_only_mode'):
            ws[f'A{row}'] = "Note: This report shows sales data and stock quantities for ciment items"
        row += 2
        
        # Headers (new structure: site, total sales, individual items, row total qty, stock quantity)
        headers = [
            'Site Name',
            'Total Sales (All Items)'
        ]
        
        # Add individual ciment item sales columns
        if metadata.get('ciment_items'):
            for item in metadata['ciment_items']:
                # Create header: "ITEM_CODE - DESCR1"
                item_code = item['ITEM']
                item_name = item['DESCR1']
                
                # Truncate item name if too long for Excel column width
                if len(item_name) > 20:
                    item_name = item_name[:17] + "..."
                
                headers.append(f"{item_code} - {item_name}")
        
        # Add summary columns at the end
        headers.extend([
            'Row Total (Qty)', 
            'Total Ciment Stock'
        ])
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, site_data in enumerate(report_data, row + 1):
            col_idx = 1
            
            # Site name
            ws.cell(row=row_idx, column=col_idx, value=site_data['SITE_NAME']).border = border
            col_idx += 1
            
            # Total sales (all items)
            ws.cell(row=row_idx, column=col_idx, value=site_data['TOTAL_SALES']).border = border
            col_idx += 1
            
            # Individual item sales
            row_total_qty = 0
            if metadata.get('ciment_items'):
                for item in metadata['ciment_items']:
                    item_code = item['ITEM']
                    sales_qty = site_data.get('CIMENT_SALES_BY_ITEM', {}).get(item_code, 0)
                    ws.cell(row=row_idx, column=col_idx, value=sales_qty).border = border
                    row_total_qty += sales_qty
                    col_idx += 1
            
            # Row total (quantity)
            ws.cell(row=row_idx, column=col_idx, value=row_total_qty).border = border
            col_idx += 1
            
            # Stock quantity
            ws.cell(row=row_idx, column=col_idx, value=site_data['TOTAL_CIMENT_STOCK']).border = border
        
        # Auto-adjust column widths (adjusted for headers with spaces)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Use appropriate width limits for headers with spaces
            adjusted_width = min(max_length + 2, 25)  # Increased from 20 to 25 for readability
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['ciment_report']
        
        if from_date and to_date:
            from_date_str = from_date.replace('-', '')
            to_date_str = to_date.replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-sales-report', methods=['POST'])
def api_export_sales_report():
    """Export sales report to Excel with proper formatting"""
    try:
        data = request.get_json()
        site_type = data.get('site_type')
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        report_date = data.get('report_date')
        
        dataframes = get_dataframes()
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the sales report data by calling the API function directly
        from routes.api_routes import api_sales_report
        from flask import current_app
        
        with current_app.test_request_context('/api/sales-report', method='POST', json=data):
            response = api_sales_report()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        site_type_name = metadata.get('site_type_name', 'Unknown')
        title = f"Sales Report - {site_type_name} Sites"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        elif from_date:
            title += f" - From: {from_date}"
        elif to_date:
            title += f" - To: {to_date}"
        
        # Write title
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Site Type: {site_type_name}"
        row += 1
        ws[f'A{row}'] = f"Total Sites: {len(report_data)}"
        row += 1
        ws[f'A{row}'] = f"Total Sales Amount: ${metadata.get('total_sales_amount', 0):,.2f} USD"
        row += 1
        ws[f'A{row}'] = f"Total Discount Amount: ${metadata.get('total_discount_amount', 0):,.2f} USD"
        row += 1
        ws[f'A{row}'] = f"Note: Sales from INVOICE.NET, Discount = SUBTOTAL-(NET+VAT+OTHER), Cumulative from ITEMS table"
        row += 1
        ws[f'A{row}'] = f"Filter: FTYPE=1 and SID starting with 530{'1' if site_type == 'kinshasa' else '2'} ({site_type_name})"
        row += 2
        
        # Headers
        headers = [
            'Site Name',
            'Sales Amount (USD)',
            'Discount Amount (USD)',
            'Cumulative Sales (USD)'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, site_data in enumerate(report_data, row + 1):
            ws.cell(row=row_idx, column=1, value=site_data['SITE_NAME']).border = border
            ws.cell(row=row_idx, column=2, value=site_data['SALES_AMOUNT']).border = border
            ws.cell(row=row_idx, column=3, value=site_data.get('DISCOUNT_AMOUNT', 0)).border = border
            ws.cell(row=row_idx, column=4, value=site_data['CUMULATIVE_SALES']).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['sales_report', site_type]
        
        selected_date = metadata.get('selected_date')
        if selected_date:
            date_str = selected_date.replace('-', '')
            filename_parts.append(date_str)
        elif report_date:
            date_str = report_date.replace('-', '')
            filename_parts.append(date_str)
        elif from_date and to_date:
            from_date_str = from_date.replace('-', '')
            to_date_str = to_date.replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-sales-by-item', methods=['POST'])
def api_export_sales_by_item():
    """Export sales by item report to Excel with proper formatting"""
    try:
        data = request.get_json()
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        site_type = data.get('site_type')
        
        dataframes = get_dataframes()
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the sales by item report data by calling the API function directly
        from routes.api_routes import api_sales_by_item_report
        from flask import current_app
        
        with current_app.test_request_context('/api/sales-by-item-report', method='POST', json=data):
            response = api_sales_by_item_report()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales by Item Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        site_type_name = metadata.get('site_type_name', 'All Sites')
        title = f"Sales by Item Report - {site_type_name}"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        
        # Write title
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Site Type: {site_type_name}"
        row += 1
        ws[f'A{row}'] = f"Period: {from_date} to {to_date}"
        row += 1
        ws[f'A{row}'] = f"Total Items: {metadata.get('total_items', 0):,}"
        row += 1
        ws[f'A{row}'] = f"Total Quantity Sold: {metadata.get('total_qty_sold', 0):,.0f}"
        row += 1
        ws[f'A{row}'] = f"Total Sales Amount: ${metadata.get('total_sales_amount', 0):,.2f} USD"
        row += 1
        ws[f'A{row}'] = f"Total Discount Amount: ${metadata.get('total_discount', 0):,.2f} USD"
        row += 1
        ws[f'A{row}'] = f"Data Source: {metadata.get('data_source', 'Sales details for quantities, inventory items for prices')}"
        row += 2
        
        # Headers
        headers = [
            'Item Code',
            'Item Name',
            'Category',
            'Prix (USD)',
            'Qty Sold',
            'Discount (USD)',
            'Total Sales (USD)'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, item_data in enumerate(report_data, row + 1):
            ws.cell(row=row_idx, column=1, value=item_data['ITEM_CODE']).border = border
            ws.cell(row=row_idx, column=2, value=item_data['ITEM_NAME']).border = border
            ws.cell(row=row_idx, column=3, value=item_data['CATEGORY']).border = border
            ws.cell(row=row_idx, column=4, value=item_data['PRIX']).border = border
            ws.cell(row=row_idx, column=5, value=int(item_data['QTY_SOLD'])).border = border
            ws.cell(row=row_idx, column=6, value=item_data['DISCOUNT']).border = border
            ws.cell(row=row_idx, column=7, value=item_data['TOTAL_SALES']).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['sales_by_item']
        
        if site_type:
            filename_parts.append(site_type)
        
        if from_date and to_date:
            from_date_str = from_date.replace('-', '')
            to_date_str = to_date.replace('-', '')
            if from_date == to_date:
                filename_parts.append(from_date_str)
            else:
                filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-bureau-client-report', methods=['POST'])
def api_export_bureau_client_report():
    """Export Kinshasa Bureau Client report to Excel with proper formatting"""
    try:
        data = request.get_json()
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        site_sidno = data.get('site_sidno')
        
        dataframes = get_dataframes()
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the bureau client report data by calling the API function directly
        from routes.api_routes import api_kinshasa_bureau_client_report
        from flask import current_app
        
        with current_app.test_request_context('/api/kinshasa-bureau-client-report', method='POST', json=data):
            response = api_kinshasa_bureau_client_report()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bureau Client Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        title = "Kinshasa Sales Bureau Client Report"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        if site_sidno:
            site_names = {
                '3700002': 'Kinshasa',
                '3700004': 'Depot Kinshasa',
                '3700003': 'Interieur'
            }
            site_name = site_names.get(str(site_sidno), f'SIDNO {site_sidno}')
            title += f" - {site_name}"
        
        # Write title
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Total Clients: {len(report_data)}"
        row += 1
        filter_text = "SID starting with 4112 (Office Clients)"
        if site_sidno:
            site_names = {
                '3700002': 'Kinshasa',
                '3700004': 'Depot Kinshasa',
                '3700003': 'Interieur'
            }
            site_name = site_names.get(str(site_sidno), f'SIDNO {site_sidno}')
            filter_text += f", SIDNO = {site_sidno} ({site_name})"
        ws[f'A{row}'] = f"Filter: {filter_text}"
        row += 1
        ws[f'A{row}'] = f"Data Source: ITEMS table (sales_details) joined with ALLSTOCK table (sites) for SIDNO filtering"
        row += 1
        ws[f'A{row}'] = f"Period: {from_date} to {to_date}"
        row += 1
        ws[f'A{row}'] = f"Total Sales Qty: {metadata.get('total_sales_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Returns Qty: {metadata.get('total_returns_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Net Qty: {metadata.get('total_net_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Invoices: {metadata.get('total_invoices', 0):,}"
        row += 1
        ws[f'A{row}'] = f"Total Quantity in USD: ${metadata.get('total_quantity_usd', 0):,.2f}"
        row += 2
        
        # Headers
        headers = [
            'Client Name',
            'Client ID (SID)',
            'Number of Invoices',
            'Sales Qty (FTYPE=1)',
            'Returns Qty (FTYPE=2)',
            'Total (Net)',
            'Quantity in USD'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, client_data in enumerate(report_data, row + 1):
            ws.cell(row=row_idx, column=1, value=client_data.get('CLIENT_NAME', 'Unknown')).border = border
            ws.cell(row=row_idx, column=2, value=client_data.get('SID', '')).border = border
            ws.cell(row=row_idx, column=3, value=client_data.get('NUM_INVOICES', 0)).border = border
            ws.cell(row=row_idx, column=4, value=client_data.get('SALES_QTY', 0)).border = border
            ws.cell(row=row_idx, column=5, value=client_data.get('RETURNS_QTY', 0)).border = border
            ws.cell(row=row_idx, column=6, value=client_data.get('TOTAL_QTY', 0)).border = border
            ws.cell(row=row_idx, column=7, value=client_data.get('QUANTITY_USD', 0)).border = border
        
        # Add totals row
        totals_row = row + len(report_data) + 1
        ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=totals_row, column=1).border = border
        ws.cell(row=totals_row, column=2, value="").border = border
        ws.cell(row=totals_row, column=3, value=metadata.get('total_invoices', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=3).border = border
        ws.cell(row=totals_row, column=4, value=metadata.get('total_sales_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=4).border = border
        ws.cell(row=totals_row, column=5, value=metadata.get('total_returns_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=5).border = border
        ws.cell(row=totals_row, column=6, value=metadata.get('total_net_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=6).border = border
        ws.cell(row=totals_row, column=7, value=metadata.get('total_quantity_usd', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=7).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['kinshasa_bureau_client_report']
        
        if site_sidno:
            site_names = {
                '3700002': 'kinshasa',
                '3700004': 'depot_kinshasa',
                '3700003': 'interieur'
            }
            site_name = site_names.get(str(site_sidno), f'sidno_{site_sidno}')
            filename_parts.append(site_name)
        
        if from_date and to_date:
            from_date_str = from_date.replace('-', '')
            to_date_str = to_date.replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-bureau-items-report', methods=['POST'])
def api_export_bureau_items_report():
    """Export Kinshasa Bureau Items report to Excel with proper formatting"""
    try:
        data = request.get_json()
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        top_n = data.get('top_n', 50)
        
        dataframes = get_dataframes()
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the bureau items report data by calling the API function directly
        from routes.api_routes import api_kinshasa_bureau_items_report
        from flask import current_app
        
        with current_app.test_request_context('/api/kinshasa-bureau-items-report', method='POST', json=data):
            response = api_kinshasa_bureau_items_report()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bureau Items Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        title = "Kinshasa Sales Bureau Top Items Report"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        
        # Write title
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Total Items: {len(report_data)}"
        row += 1
        ws[f'A{row}'] = f"Top N: {top_n if top_n != '0' else 'All Items'}"
        row += 1
        ws[f'A{row}'] = f"Filter: SID starting with 4112 (Office Clients)"
        row += 1
        ws[f'A{row}'] = f"Data Source: ITEMS table (sales_details)"
        row += 1
        ws[f'A{row}'] = f"Period: {from_date} to {to_date}"
        row += 1
        ws[f'A{row}'] = f"Total Sales Qty: {metadata.get('total_sales_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Returns Qty: {metadata.get('total_returns_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Net Qty: {metadata.get('total_net_qty', 0):,.2f}"
        row += 2
        
        # Headers
        headers = [
            'Item Code',
            'Item Name (DESCR1)',
            'Category',
            'Sales Qty (FTYPE=1)',
            'Returns Qty (FTYPE=2)',
            'Total (Net)'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, item_data in enumerate(report_data, row + 1):
            ws.cell(row=row_idx, column=1, value=item_data.get('ITEM_CODE', '')).border = border
            ws.cell(row=row_idx, column=2, value=item_data.get('ITEM_NAME', 'Unknown')).border = border
            ws.cell(row=row_idx, column=3, value=item_data.get('CATEGORY', 'Unknown Category')).border = border
            ws.cell(row=row_idx, column=4, value=item_data.get('SALES_QTY', 0)).border = border
            ws.cell(row=row_idx, column=5, value=item_data.get('RETURNS_QTY', 0)).border = border
            ws.cell(row=row_idx, column=6, value=item_data.get('TOTAL_QTY', 0)).border = border
        
        # Add totals row
        totals_row = row + len(report_data) + 1
        ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=totals_row, column=1).border = border
        ws.cell(row=totals_row, column=2, value="").border = border
        ws.cell(row=totals_row, column=3, value="").border = border
        ws.cell(row=totals_row, column=4, value=metadata.get('total_sales_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=4).border = border
        ws.cell(row=totals_row, column=5, value=metadata.get('total_returns_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=5).border = border
        ws.cell(row=totals_row, column=6, value=metadata.get('total_net_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=6).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['kinshasa_bureau_items_report']
        
        if from_date and to_date:
            from_date_str = from_date.replace('-', '')
            to_date_str = to_date.replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/export-item-clients-report', methods=['POST'])
def api_export_item_clients_report():
    """Export Kinshasa Bureau Item Clients report to Excel with proper formatting"""
    try:
        data = request.get_json()
        item_code = data.get('item_code')
        from_date = data.get('from_date')
        to_date = data.get('to_date')
        
        dataframes = get_dataframes()
        if not dataframes:
            return jsonify({'error': 'No data loaded. Please load dataframes first.'}), 400
        
        # Get the item clients report data by calling the API function directly
        from routes.api_routes import api_kinshasa_bureau_item_clients
        from flask import current_app
        
        with current_app.test_request_context('/api/kinshasa-bureau-item-clients', method='POST', json=data):
            response = api_kinshasa_bureau_item_clients()
            if hasattr(response, 'status_code') and response.status_code != 200:
                return response
            
            response_data = response.get_json() if hasattr(response, 'get_json') else response
        
        if 'error' in response_data:
            return jsonify(response_data), 400
        
        report_data = response_data['data']
        metadata = response_data['metadata']
        
        if not report_data:
            return jsonify({'error': 'No data to export'}), 404
        
        # Create Excel workbook
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Item Clients Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title and metadata
        title = f"Item Clients Report - {metadata.get('item_code', 'Unknown')}"
        if from_date and to_date:
            title += f" - Period: {from_date} to {to_date}"
        
        # Write title
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Write metadata
        row = 3
        ws[f'A{row}'] = f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
        row += 1
        ws[f'A{row}'] = f"Item Code: {metadata.get('item_code', 'Unknown')}"
        row += 1
        ws[f'A{row}'] = f"Item Name: {metadata.get('item_name', 'Unknown')}"
        row += 1
        ws[f'A{row}'] = f"Total Clients: {len(report_data)}"
        row += 1
        ws[f'A{row}'] = f"Filter: SID starting with 4112 (Office Clients)"
        row += 1
        ws[f'A{row}'] = f"Data Source: ITEMS table (sales_details)"
        row += 1
        ws[f'A{row}'] = f"Period: {from_date} to {to_date}"
        row += 1
        ws[f'A{row}'] = f"Total Sales Qty: {metadata.get('total_sales_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Returns Qty: {metadata.get('total_returns_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Net Qty: {metadata.get('total_net_qty', 0):,.2f}"
        row += 1
        ws[f'A{row}'] = f"Total Invoices: {metadata.get('total_invoices', 0):,}"
        row += 1
        ws[f'A{row}'] = f"Total Quantity in USD: ${metadata.get('total_quantity_usd', 0):,.2f}"
        row += 2
        
        # Headers
        headers = [
            'Client Name',
            'Client ID (SID)',
            'Number of Invoices',
            'Sales Qty (FTYPE=1)',
            'Returns Qty (FTYPE=2)',
            'Total (Net)',
            'Quantity in USD'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row_idx, client_data in enumerate(report_data, row + 1):
            ws.cell(row=row_idx, column=1, value=client_data.get('CLIENT_NAME', 'Unknown')).border = border
            ws.cell(row=row_idx, column=2, value=client_data.get('SID', '')).border = border
            ws.cell(row=row_idx, column=3, value=client_data.get('NUM_INVOICES', 0)).border = border
            ws.cell(row=row_idx, column=4, value=client_data.get('SALES_QTY', 0)).border = border
            ws.cell(row=row_idx, column=5, value=client_data.get('RETURNS_QTY', 0)).border = border
            ws.cell(row=row_idx, column=6, value=client_data.get('TOTAL_QTY', 0)).border = border
            ws.cell(row=row_idx, column=7, value=client_data.get('QUANTITY_USD', 0)).border = border
        
        # Add totals row
        totals_row = row + len(report_data) + 1
        ws.cell(row=totals_row, column=1, value="TOTALS").font = Font(bold=True)
        ws.cell(row=totals_row, column=1).border = border
        ws.cell(row=totals_row, column=2, value="").border = border
        ws.cell(row=totals_row, column=3, value=metadata.get('total_invoices', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=3).border = border
        ws.cell(row=totals_row, column=4, value=metadata.get('total_sales_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=4).border = border
        ws.cell(row=totals_row, column=5, value=metadata.get('total_returns_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=5).border = border
        ws.cell(row=totals_row, column=6, value=metadata.get('total_net_qty', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=6).border = border
        ws.cell(row=totals_row, column=7, value=metadata.get('total_quantity_usd', 0)).font = Font(bold=True)
        ws.cell(row=totals_row, column=7).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['item_clients_report', item_code]
        
        if from_date and to_date:
            from_date_str = from_date.replace('-', '')
            to_date_str = to_date.replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        filename = '_'.join(filename_parts) + '.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500