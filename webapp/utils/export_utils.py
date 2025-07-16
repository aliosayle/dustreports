"""Export utilities for generating Excel and PDF reports"""

import io
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


class ExcelExporter:
    """Handles Excel export functionality"""
    
    @staticmethod
    def create_stock_autonomy_report(report_data, filters):
        """Create Excel export for stock autonomy report"""
        
        if not report_data:
            raise ValueError('No data to export')
        
        df = pd.DataFrame(report_data)
        
        # Build title with filters
        title = "Stock Autonomy Report"
        filter_info = ExcelExporter._build_autonomy_filter_info(filters)
        
        if filter_info:
            title += f" - {' | '.join(filter_info)}"
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock Autonomy', index=False, startrow=2)
            
            # Style the workbook
            ExcelExporter._style_workbook(writer, title, df, 'Stock Autonomy')
        
        output.seek(0)
        
        # Generate filename
        filename = ExcelExporter._generate_autonomy_filename(filters)
        
        return output, filename
    
    @staticmethod
    def create_stock_by_site_report(report_data, filters):
        """Create Excel export for stock by site report"""
        
        if not report_data:
            raise ValueError('No data to export')
        
        df = pd.DataFrame(report_data)
        
        # Build title with filters
        title = "Stock by Site Report"
        filter_info = ExcelExporter._build_stock_by_site_filter_info(filters)
        
        if filter_info:
            title += f" - {' | '.join(filter_info)}"
        
        # Create Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock by Site', index=False, startrow=2)
            
            # Style the workbook
            ExcelExporter._style_workbook(writer, title, df, 'Stock by Site')
        
        output.seek(0)
        
        # Generate filename
        filename = ExcelExporter._generate_stock_by_site_filename(filters)
        
        return output, filename
    
    @staticmethod
    def _build_autonomy_filter_info(filters):
        """Build filter information for autonomy report"""
        filter_info = []
        
        if filters.get('item_code'):
            filter_info.append(f"Item: {filters['item_code']}")
        if filters.get('site_code'):
            filter_info.append(f"Site: {filters['site_code']}")
        if filters.get('from_date'):
            filter_info.append(f"From: {filters['from_date']}")
        if filters.get('to_date'):
            filter_info.append(f"To: {filters['to_date']}")
        
        return filter_info
    
    @staticmethod
    def _build_stock_by_site_filter_info(filters):
        """Build filter information for stock by site report"""
        filter_info = []
        
        if filters.get('item_code'):
            filter_info.append(f"Item: {filters['item_code']}")
        if filters.get('site_codes') and len(filters['site_codes']) > 0:
            site_codes = filters['site_codes']
            if len(site_codes) == 1:
                filter_info.append(f"Site: {site_codes[0]}")
            else:
                filter_info.append(f"Sites: {len(site_codes)} selected")
        if filters.get('category_id'):
            filter_info.append(f"Category: {filters['category_id']}")
        if filters.get('as_of_date'):
            filter_info.append(f"As of Date: {filters['as_of_date']}")
        
        return filter_info
    
    @staticmethod
    def _style_workbook(writer, title, df, sheet_name):
        """Apply styling to Excel workbook"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Add title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(size=14, bold=True)
        
        # Merge cells for title
        last_column = len(df.columns)
        last_column_letter = get_column_letter(last_column)
        
        if last_column > 1:
            worksheet.merge_cells(f'A1:{last_column_letter}1')
        else:
            worksheet.merge_cells('A1:A1')
        
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Style header row
        header_fill = PatternFill(start_color='1f4e79', end_color='1f4e79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _generate_autonomy_filename(filters):
        """Generate filename for autonomy report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['stock_autonomy_report']
        
        if filters.get('site_code'):
            site_code = str(filters['site_code']).replace(' ', '_').replace('/', '_')
            filename_parts.append(f'site_{site_code}')
        
        if filters.get('item_code'):
            item_code = str(filters['item_code']).replace(' ', '_').replace('/', '_')
            filename_parts.append(f'item_{item_code}')
        
        if filters.get('from_date') and filters.get('to_date'):
            from_date_str = filters['from_date'].replace('-', '')
            to_date_str = filters['to_date'].replace('-', '')
            filename_parts.append(f'{from_date_str}_to_{to_date_str}')
        
        filename_parts.append(timestamp)
        return '_'.join(filename_parts) + '.xlsx'
    
    @staticmethod
    def _generate_stock_by_site_filename(filters):
        """Generate filename for stock by site report"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_parts = ['stock_by_site_report']
        
        if filters.get('site_codes') and len(filters['site_codes']) > 0:
            filename_parts.append(f'sites_{len(filters["site_codes"])}')
        
        if filters.get('category_id'):
            category_id = str(filters['category_id']).replace(' ', '_').replace('/', '_')
            filename_parts.append(f'category_{category_id}')
        
        if filters.get('as_of_date'):
            as_of_date_str = filters['as_of_date'].replace('-', '')
            filename_parts.append(f'as_of_{as_of_date_str}')
        
        filename_parts.append(timestamp)
        return '_'.join(filename_parts) + '.xlsx'


class PDFExporter:
    """Handles PDF export functionality"""
    
    @staticmethod
    def create_report(report_data, title, filters=None):
        """Create a PDF report"""
        
        if not report_data:
            raise ValueError('No data to export')
        
        df = pd.DataFrame(report_data)
        
        # Create PDF in memory
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        full_title = title
        if filters:
            filter_info = []
            for key, value in filters.items():
                if value:
                    filter_info.append(f"{key.replace('_', ' ').title()}: {value}")
            if filter_info:
                full_title += f"<br/><font size=12>Filters: {' | '.join(filter_info)}</font>"
        
        elements.append(Paragraph(full_title, title_style))
        elements.append(Spacer(1, 12))
        
        # Convert DataFrame to table data
        data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        return output
